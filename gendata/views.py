from django.http import HttpResponse
from django.shortcuts import redirect, render
import openai
import json
from os import environ
import openpyxl

OPENAI_API_KEY = environ.get("OPENAI_API_KEY")
# Create your views here.
def index(request):
    return render(request, 'index.html')

def prompt(request):
    return render(request, 'prompt.html')

def rsa(request):

    type = request.POST.get('type')
    
    product = request.POST.get('rsa-product')
    cta = request.POST.get('rsa-cta')
    finalUrl = request.POST.get('rsa-finalurl')
    tone = request.POST['rsa-tone']
    audience = request.POST['rsa-audience']
    case = request.POST.get('rsa-case')


    # if type == 'rsa':
    #     target = request.POST.get('rsa_target')

    main_prompt = "Create an RSA with unique headlines and descriptions,\
        include the CTAs. Stick to the character limits and rules here: \
        Maximum of 15 headlines per ad. Each headline can have up to 30 \
        characters. Provide only 2 descriptions. Each description\
        can have up to 90 characters. CTA = call to action. RSA = Responsive\
        Search Ad. Provide response in the JSON format with the keys: headline1, headline2, ..., description1, description2.\n"
    prompt = main_prompt + "Product:" + product + "Tone:" + tone + "CTA:" + cta + "Target audience:" + audience + "Case:" + case
    response = get_response(prompt)
    # else:
    #     urls = request.POST.get('sl_urls')

    #     main_prompt = "Create Google Ads sitelink extensions with unique headlines \
    #         and descriptions using the following final URLs. Include the final URLs \
    #         in your response. Stick to the format and character limits.\
    #         Link text: 25 characters\
    #         Description Line 1: 35 characters\
    #         Description Line 2: 35 characters"
    #     prompt = main_prompt + "Product:" + product + "Tone:" + tone + "Final URLs:" + urls
    #     response = get_response(prompt)
    response = json.loads(response)
    response['finalUrl'] = finalUrl
    print((response))
    # response = json.dumps(response)
    # print(type(response))
    return render(request, 'rsa.html', {'response':response})


def get_response(prompt):
    messages = []
    messages.append({"role": "system", "content": "Your name is AI Stuff. You are a helpful assistant."})

    question = {}
    question['role'] = 'user'
    question['content'] = prompt
    messages.append(question)
    response = openai.ChatCompletion.create(model="gpt-3.5-turbo",messages=messages)
    
    response = response['choices'][0]['message']['content']
    # return HttpResponse(
    #     json.dumps(result),
    #     content_type="application/json"
    # )
    return response
    # return HttpResponse(response)

def generate_sitelink(request):
    pass

def download(request):
    print("Download")
    response = HttpResponse(content_type='application/ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="rsa.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    # Write the data to the Excel file
    i = 1
    global data
    for key, value in data.items():
        title_cell = worksheet.cell(row=1, column=i)
        title_cell.value = key
        
        rsa_cell = worksheet.cell(row=2, column=i)
        rsa_cell.value = value
        
        i += 1

    # Save the Excel file to the response
    workbook.save(response)
    
    return response